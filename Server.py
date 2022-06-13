from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from visual_server import *
import sys
import socket
from time import sleep
import json
import os
import pandas as pd
from openpyxl import load_workbook


class Process:
    def __init__(self):
        self.pid = 1111
        self.name = "process.exe"
        self.status = "running"
        self.time = 0
        self.rss = 0

    def cout(self):
        print(self.pid, self.name, self.status, self.time, round(self.rss / (1024 * 1024), 3), "MB")


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        # Визуальная составляющая
        QtWidgets.QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Некоторые переменные
        self.worker = False
        self.rowPosition = 0
        self.HOST = str(self.ui.lineEdit_HOST.text())
        self.PORT = int(self.ui.lineEdit_PORT.text())
        self.path_to_save_excel = str(self.ui.lineEdit_save_to_excel.text())   # "D:\\"

        # Style clicked menu button
        for w in self.ui.frame_3.findChildren(QtWidgets.QPushButton):
            w.clicked.connect(self.applyButtonStyle)   # Add click event listener

        # Функционал кнопок
        self.ui.pushButton_connect.clicked.connect(lambda: self.applyButtonStyle)

        self.show()


    def applyButtonStyle(self):
        # если Режим Приема был выкл. (False), а стал вкл. (True), то запускай функцию обработки
        if self.worker == True:
            print("Сервер: ВЫКЛ \n")
            self.sender().setStyleSheet("border-bottom: none;")
            self.worker = False
            self.my_thread.disconnect()

        else:
            print("Сервер: ВКЛ")
            self.sender().setStyleSheet("border-bottom: 2px solid;")
            self.worker = True

            # Функционал потока и сигнала
            self.my_thread = MyThread(self.worker, self.HOST, self.PORT, self.path_to_save_excel)
            self.my_thread.my_signal.connect(self.processes)   # Подключение сигнала в поток
            self.my_thread.start()
            self.ui.pushButton_2.clicked.connect(self.my_thread.activate)
        return


# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------Get running processes--------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
    def processes(self, stroka):
        stroka = stroka.split("!@!")
        count, pid, name, status = stroka[0], stroka[1], stroka[2], stroka[3]
        time, rss, self.rowPosition = stroka[4], stroka[5], int(stroka[6])

        # Create new Raw
        table_rowPosition = self.ui.tableWidget.rowCount()
        if self.rowPosition == table_rowPosition:
            self.ui.tableWidget.insertRow(self.rowPosition)
        try:
            # Create widget
            self.create_table_widget(0, pid, "tableWidget")
            self.create_table_widget(1, name, "tableWidget")
            self.create_table_widget(2, status, "tableWidget")
            self.create_table_widget(3, time, "tableWidget")
            self.create_table_widget(4, rss + " MB", "tableWidget")

        except Exception as e:
            print(e, '\n')

        if int(self.rowPosition) == int(count)-1:   # Таблица полностью заполненна, обнули флаг строк для след итерации
            self.rowPosition = 0

        self.ui.lineEdit_SEARCH.textChanged.connect(self.findName)  # activity_search.field
# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------Function that CREATE table widget-------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
    def create_table_widget(self, columnPosition, text, tableName):
        qtablewidgetitem = QtWidgets.QTableWidgetItem()

        # use method getattr()
        getattr(self.ui, tableName).setItem(self.rowPosition, columnPosition, qtablewidgetitem)
        qtablewidgetitem = getattr(self.ui, tableName).item(self.rowPosition, columnPosition)

        qtablewidgetitem.setText(str(text))
# ----------------------------------------------------------------------------------------------------------------------
# ---------------------------------- searh activity table --------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
    def findName(self):
        name = self.ui.lineEdit_SEARCH.text().lower()
        for row in range(self.ui.tableWidget.rowCount()):
            item = self.ui.tableWidget.item(row, 1)
            # if the search is "not" in the item`s text "do not hide" the row
            self.ui.tableWidget.setRowHidden(row, name not in item.text().lower())








class MyThread(QThread):
    my_signal = pyqtSignal(str)  # 1

    def __init__(self, worker, HOST, PORT, path_to_save_excel):
        # Некоторые переменные
        super(MyThread, self).__init__()
        self.count = 0
        self.rowPosition = 0
        self.worker = worker   # Показывает состояние сервера
        self.HOST = HOST
        self.PORT = PORT
        self.path = path_to_save_excel
        self.activator = False      # Отвечает за момент активации запроса на получение данных от клиента
                                    # Флаг для определения состояния Активатора (блок общения)


    def disconnect(self):
        self.worker = False   # Флаг для определения состояния Сервера (блок соединения)
        self.activator = False


    def activate(self):
        self.activator = True


    def connect(self):
        server_socket = socket.socket()  # создание сокета
        server_socket.bind((self.HOST, self.PORT))  # определение хоста и порта
        server_socket.listen(1)  # режим прослушивания (1 - макс кол подключений в очереди)
        server_socket, addr = server_socket.accept()  # можем принять подключение (кортеж с двумя элементами: новый сокет и адрес клиента)
        # self.activator = True
        return server_socket, addr


    def run(self):
        # Некоторые переменные
        self.rowPosition = 0
        pid_list = []
        name_list = []
        status_list = []
        time_list = []
        rss_list = []
        node = ""
        login = ""
        page_name = ""
        try:
            server_socket, addr = self.connect()
            if self.worker == False:   # Сервер выключен
                server_socket.close()
                return
            print('   connected:', addr)

            while True:
                print(self.activator)
                sleep(1)   # Проверяет наличие команды на активацию запроса данных

                if self.activator == True:
                    print("   Активировать запрос на передачу данных")
                    message = bytes(json.dumps({"start": "Запрашиваю передачу данных"}), encoding="utf-8")
                    server_socket.sendall(message)  # отправка сообщения на сервер
                    sleep(0.1)
                    print('   Запрос отправлен')
                    del message

                    while True:
                        data = server_socket.recv(1024).decode("utf-8")
                        if data:
                            message = json.loads(data)
                            print(message["pid"])

                            # Формирование Данных для Excel-таблицы
                            count = int(message["count"])
                            print(count)
                            pid_list.append(int(message["pid"]))
                            name_list.append(message["name"])
                            status_list.append(message["status"])
                            time_list.append(message["my_time"])
                            rss_list.append(message["rss"])
                            if self.rowPosition == 1:
                                node = message["node"]
                                login = message["login"]
                                page_name = message["page_name"]

                            # Формирование Данных для возврата в MainWindow
                            answer = str(message["count"]) + "!@!"
                            answer += message["pid"] + "!@!"
                            answer += message["name"] + "!@!"
                            answer += message["status"] + "!@!"
                            answer += message["my_time"] + "!@!"
                            answer += str(message["rss"]) + "!@!" + str(self.rowPosition)
                            self.my_signal.emit(answer)   # Возврат Данных в MainWindow

                            self.rowPosition += 1
                            if self.rowPosition == count:   # получена последняя строка
                                print("   Сервер: прием данных окончен, отправляю повторный запрос")
                                message = bytes(json.dumps({"finish": "   Сервер: Данные принял, благодарю за работу"}), encoding="utf-8")
                                server_socket.sendall(message)  # отправка сообщения на сервер

                                # Запись в xlsx-таблицу
                                self.save_to_excell(pid_list, name_list, status_list, time_list, rss_list, node, login, page_name)
                                break
                            # # message = json.loads(data)
                            # print(message["finish"])
                            #
                            # # break
                        else:
                            print("   Сервер: прием данных окончен")
                            break
                    break

            sleep(0.1)
            server_socket.close()

            if self.worker == True:   # Сервер работает
                print("   rerun")
                self.activator = False
                self.run()

        except Exception as e:
            print(e, '\n')

    # ------------------------------------------------------------------------------------------------------------------
    # ----------------------------Get running processes-----------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def save_to_excell(self, pid_list, name_list, status_list, time_list, rss_list, node, login, page_name):
        # Формирование Имени Файла
        file_name = "!!!BD!!!{}.xlsx".format(login + "!!!" + node + "!!!")
        full_file_name = self.path + "\\" + file_name

        # Проверка Существования Файла
        file_list = []
        for _, _, files in os.walk(self.path):
            for file in files:
                if file[-4:] == "xlsx":
                    file_list.append(file)
            break
        # Есть такой?
        if file_name in file_list:
            try:
                print("   Файл есть, все норм.")
            except Exception:
                print("   Ошибка файла! ", Exception)
        # Нет, создаем новый
        else:
            print("   Создаем новый файл.")
            df = pd.DataFrame({'Testing': [6000000, 600000]})
            df.to_excel(full_file_name)
            del df


        # Просмотр параметров книги
        wb = load_workbook(full_file_name)  # Вся книга == Список Листов (страниц)
        # sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])  # Выбор конкретной страницы   (В данном случае не нужно)
        print("   ", file_name, wb.get_sheet_names())

        cols = ["Pid", "Name", "Status", "Time", "Rss (Mb)"]
        new_data = []
        for iterator, pid in enumerate(pid_list):

            # # Адаптируем "Время" в нормальный для записи форма   (Заюзать при необходимости)
            # timming = time_list[iterator]
            # split_list = " ", ":"
            # for split_point in split_list:
            #     new_timming = ""
            #     for time_point in timming.split(split_point):
            #         new_timming += time_point + "-"
            #     timming = new_timming
            # print(timming)

            t = (pid, name_list[iterator], status_list[iterator], time_list[iterator], rss_list[iterator])
            new_data.append(t)   # Каждый элемент t -строка, а каждый *_list - столбец
        df = pd.DataFrame(new_data, index=range(1, len(pid_list)+1), columns=cols)

        # Формируем Имя Страницы
        split_symbol = [" ", ":"]
        for ss in split_symbol:
            str_name = ""
            for i in page_name.split(ss):
                str_name += "_" + i
            page_name = str_name
        page_name = "!BD!_" + page_name[2:]

        # Формируем Страницу
        from openpyxl.utils.dataframe import dataframe_to_rows
        print("   Записи в *xlsx")

        wb.remove(wb[wb.get_sheet_names()[0]])  # Удаление первого листа
        ws = wb.create_sheet(page_name, 1)  # Создание нового листа
        for r in dataframe_to_rows(df, index=False, header=True):  # Построчная запись
            ws.append(r)
        wb.save(full_file_name)  # Сохранение изменений документа



# ----------------------------------------------------------------------------------------------------------------------
# ------------------------------Запуск визуалки-------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec_())